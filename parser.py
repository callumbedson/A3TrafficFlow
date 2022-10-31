from openpyxl import load_workbook, Workbook
import math

#Open excel file (workbook) to read
newWorkbook = Workbook()
workbook = load_workbook(filename="C:/Users/bedsoc/Downloads/Scats Data October 2006.xlsx", data_only=True)

#Read from the dat sheet
sheet = workbook["Data"]
newSheet = newWorkbook.active

#Create array for scats sites
scatsList = []

#iterate through rows
for x in range(3,sheet.max_row):
        #get the two roads, location splitting by "of"
        roads = sheet["B"+str(x)].value.lower().split(" of ")
        #remove the compass direction by removing last two letters, and strip whitespace so (eg E and NE) end consistently
        #In hindsight, the compass direction could have been used for the direct connection check below
        roads[0] = roads[0][:-2].strip()
        #new is to keep track of whether road number had been added before
        #if it has, only add the new streets
        new = True 
        #loop through added scats
        for scatsNum in scatsList:
            if sheet["A"+str(x)].value == scatsNum[0]:
                new = False
                #add any new streets
                for road in roads:
                    if road not in scatsNum[1]:
                        scatsNum[1].append(road)
        if new:
            #Convert longitude and latitude to km values before appending
            #From https://en.wikipedia.org/wiki/Latitude#Meridian_distance_on_the_ellipsoid
            #Technically curvature should be applied, but we use simple calculation for local boorondara area (~37 degrees)
            scatsList.append((sheet["A"+str(x)].value, roads ,sheet["D"+str(x)].value * 110.852, sheet["E"+str(x)].value * 96.486))

#For each scat (X) site, check other scat (y) sites
for scatsX in scatsList:
    for scatsY in scatsList:
        #If its trying to check itself, skip
        if scatsX == scatsY:
            continue
        for roadX in scatsX[1]:
            for roadY in scatsY[1]:
                #Compare roads in both scats sites
                #If any match
                if roadX == roadY:
                    #Prepare to add it
                    addIt = True
                    #Get the distance between the two
                    dist = math.sqrt((scatsX[2] - scatsY[2])**2 + (scatsX[3] - scatsY[3])**2)
                    #Also check every other scat (Z) and their roads
                    for scatsZ in scatsList:
                        for roadZ in scatsZ[1]:
                            #and again check if its itself
                            if scatsX == scatsZ:
                                continue
                            if roadZ == roadX:
                                #If any roads match beween scat x, y, and z
                                #If z is closer to x and in the same direction, do not add y (addIt = False)
                                if scatsY[2] < scatsX[2] and scatsY[3] < scatsX[3]:
                                    if scatsZ[2] < scatsX[2] and scatsZ[3] < scatsX[3] and scatsZ[2] > scatsY[2] and scatsZ[3] > scatsY[3]:
                                        print(scatsZ[0], " is closer")
                                        addIt = False
                                if scatsY[2] > scatsX[2] and scatsY[3] > scatsX[3]:
                                    if scatsZ[2] > scatsX[2] and scatsZ[3] > scatsX[3] and scatsZ[2] < scatsY[2] and scatsZ[3] < scatsY[3]:
                                        print(scatsZ[0], " is closer")
                                        addIt = False
                                if scatsY[2] < scatsX[2] and scatsY[3] > scatsX[3]:
                                    if scatsZ[2] < scatsX[2] and scatsZ[3] > scatsX[3] and scatsZ[2] > scatsY[2] and scatsZ[3] < scatsY[3]:
                                        print(scatsZ[0], " is closer")
                                        addIt = False
                                if scatsY[2] > scatsX[2] and scatsY[3] < scatsX[3]:
                                    if scatsZ[2] > scatsX[2] and scatsZ[3] < scatsX[3] and scatsZ[2] < scatsY[2] and scatsZ[3] > scatsY[3]:
                                        print(scatsZ[0], " is closer")
                                        addIt = False
                    if addIt == True:
                        #If its added, make a new row in the output excel (later we manually converted to .csv for processing speed)
                        print(scatsX, " added")
                        row = newSheet.max_row + 1
                        newSheet["A" + str(row)] = scatsX[0]
                        newSheet["B" + str(row)] = scatsY[0]
                        newSheet["C" + str(row)] = dist
                        print("dist ", dist)
    #dont recheck any lines later
    scatsList.remove(scatsX)

#save to my desktop (moved and converted later)
newWorkbook.save("C:/Users/bedsoc/Desktop/SCATSGraph.xlsx")