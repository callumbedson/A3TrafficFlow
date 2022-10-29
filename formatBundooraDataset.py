from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from datetime import datetime

newWorkbook = Workbook()
workbook = load_workbook(filename="C:/Users/jettc/OneDrive - Swinburne University/3rd Year/Semester 2/Intelligent systems/Assignment A3/Scats Data October 2006.xlsx", data_only=True)

sheet = workbook["Data"]
newSheet = newWorkbook.active

for cell in sheet['J']:
    if cell.row > 2 and cell.row < 5:

        for timeCell in (sheet["K"+str(cell.row)+":"+"DB"+str(cell.row)])[0]:
            row = newSheet.max_row + 1
            newSheet["A" + str(row)] = cell.value.replace(hour=0,minute=0).strftime("%d/%m/%Y") + " " + str(sheet[get_column_letter(timeCell.column)+"1"].value)
            print(cell.value)
            newSheet["B"+ str(row)] = timeCell.value
        print(cell.row)

newWorkbook.save("C:/Users/jettc/OneDrive - Swinburne University/3rd Year/Semester 2/Intelligent systems/Assignment A3/filteredData.xlsx")